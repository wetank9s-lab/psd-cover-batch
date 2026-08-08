# -*- coding: utf-8 -*-
"""
excel_data.py —— 统一 Excel 数据管线（Stage 4）。

Stage 4 目标：GUI（_load / _preview / run_batch）与 CLI（run / inspect）
**全部** 通过 load_excel_dataset() 这一个入口读取 Excel，不再各自直接调用
openpyxl.load_workbook / iter_rows。内部列索引统一为 0 起（A=0, B=1, ...）。

关键约定：
  - ExcelRow.values   : 保留该行的全部原始单元格（str 形式，供 Stage 4.5 扩展列使用）
  - ExcelDataset      : 加载后的不可变视图；后续只操作 ExcelRow，不持有工作簿对象
  - 物理行号 excel_row: 1 起。有表头时数据从第 2 行开始（表头自身也是第 1 行）；
                        无表头时第 1 行就是数据。
  - 电话规范化 format_phone：
      * int            -> str
      * float 且为整数 -> str(int(v))          （13800000001.0 -> '13800000001'）
      * float 非整数   -> 保留原样 str(v)       （123.45 -> '123.45'）
      * str            -> strip（**保留前导零**：'00123456' 不得变成 '123456'）
      * bool           -> '1' / '0'
      * None / 公式计算后为 None -> ''（视为空值）
  - .xls 拒绝：仅支持 .xlsx / .xlsm（不引入 xlrd）。
  - sheet 选择：第一版使用 workbook.active，并记录 sheet_name。
"""

from dataclasses import dataclass, field
from typing import Any, List, Optional, Sequence, Tuple

import os

# openpyxl 延迟导入：本模块在纯函数层也能被测试（无文件时不需要 openpyxl）。
_OPENPYXL = None


def _get_openpyxl():
    global _OPENPYXL
    if _OPENPYXL is None:
        try:
            import openpyxl
        except ImportError as e:  # pragma: no cover
            raise ExcelDataError(
                "未安装 openpyxl，请先执行：pip install openpyxl") from e
        _OPENPYXL = openpyxl
    return _OPENPYXL


# ---------------------------------------------------------------------------
# 异常
# ---------------------------------------------------------------------------
class ExcelDataError(Exception):
    """统一 Excel 数据异常。

    code 取值：
      FILE_NOT_FOUND   文件不存在
      UNSUPPORTED_TYPE 扩展名不是 .xlsx / .xlsm（如 .xls）
      WORKBOOK_BROKEN  工作簿无法打开 / 损坏
      NO_SHEET         工作簿内没有工作表
      COL_OUT_OF_RANGE 列索引越界
      NO_VALID_DATA    构建成功但没有可生成的有效数据
      COL_CONFLICT     字段列冲突（如姓名 == 电话）
    """

    def __init__(self, message: str, code: str = "GENERIC"):
        super().__init__(message)
        self.code = code


# ---------------------------------------------------------------------------
# 数据模型
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class ExcelRow:
    """从 Excel 一行解析出的业务数据。

    - excel_row: Excel 物理行号（1 起，含表头行）
    - store / name / phone / role: 已 strip 的业务字段（phone 经 format_phone）
    - values: 该行全部原始单元格的 str 形式（供 Stage 4.5 按列索引取扩展字段）
    """

    excel_row: int
    store: str
    name: str
    phone: str
    role: Optional[str] = None
    values: Tuple[str, ...] = field(default_factory=tuple)

    @property
    def excel_row_index(self) -> int:
        """0 起行索引（数据表内偏移），与 GUI 习惯的「第 1 条数据」对应。"""
        return self.excel_row - 1


@dataclass(frozen=True)
class SkippedRow:
    """被跳过的行（记录原因，便于用户核对）。"""

    excel_row: int
    reason: str


@dataclass
class ExcelDataset:
    """统一数据集（加载后只操作本对象，不持有工作簿）。"""

    path: str
    headers: List[str]                 # 表头行原值（无表头时为空列表）
    rows: List[Tuple[Any, ...]]        # 全部原始行（含表头；iter_rows 原样）
    valid_rows: List[ExcelRow]         # 通过校验的业务行
    skipped_rows: List[SkippedRow]     # 被跳过的行及原因
    stores: List[str]                  # strip、保序、去重（来自 valid_rows 的 store）
    max_columns: int                   # 最大列数（用于动态生成列下拉）
    sheet_name: str                    # 实际使用的 sheet 名

    @property
    def has_valid_data(self) -> bool:
        return bool(self.valid_rows)


# ---------------------------------------------------------------------------
# 列名工具（A..Z, AA, AB, ...）
# ---------------------------------------------------------------------------
def index_to_excel_column(index: int) -> str:
    """0 起列索引 -> Excel 列名：0->A, 25->Z, 26->AA, 27->AB ..."""
    if index < 0:
        raise ValueError(f"列索引不能为负: {index}")
    n = index
    out = []
    while True:
        out.append(chr(ord("A") + n % 26))
        n = n // 26 - 1
        if n < 0:
            break
    return "".join(reversed(out))


def excel_column_to_index(col: str) -> int:
    """Excel 列名 -> 0 起索引：'A'->0, 'Z'->25, 'AA'->26, 'ab'->26 ..."""
    s = str(col).strip().upper()
    if not s:
        raise ValueError("列名为空")
    idx = 0
    for ch in s:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"非法列名: {col!r}")
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx - 1


# ---------------------------------------------------------------------------
# 电话规范化
# ---------------------------------------------------------------------------
def format_phone(value: Any) -> str:
    """把 Excel 单元格的电话值规范成字符串。

    - int           -> str(value)
    - float 且是整数 -> str(int(value))     （如 13800000001.0 -> '13800000001'）
    - float 非整数   -> str(value)          （保留原样，避免失真）
    - str / 其他     -> str(value).strip()  （**保留前导零**）
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


# ---------------------------------------------------------------------------
# 行解析（纯函数，Stage 0 兼容）
# ---------------------------------------------------------------------------
def parse_row(
    raw: Optional[Tuple[Any, ...]],
    *,
    excel_row: int,
    col_store: int = 0,
    col_name: int = 1,
    col_phone: int = 3,
    col_role: int = 2,
    require_store: bool = False,
    require_name: bool = True,
    require_phone: bool = True,
) -> Optional[ExcelRow]:
    """把 openpyxl 的一行（tuple）解析为 ExcelRow。

    - 列索引越界时按「空值」处理；
    - 空白字符串视为空值；
    - require_store=True 时门店为空的行返回 None（跳过）；
    - name 为空（或未配置必填 phone 时 phone 为空）返回 None（跳过）；
    - 保留物理行号 excel_row。
    """
    if raw is None:
        return None
    n = len(raw)

    def cell(idx):
        return raw[idx] if 0 <= idx < n else None

    store = cell(col_store)
    name = cell(col_name)
    phone = cell(col_phone)
    role = cell(col_role) if col_role is not None and col_role >= 0 else None

    store_s = str(store).strip() if store is not None else ""
    name_s = str(name).strip() if name is not None else ""
    role_s = str(role).strip() if role is not None else ""
    phone_s = format_phone(phone)

    if require_store and not store_s:
        return None
    if require_name and not name_s:
        return None
    if require_phone and not phone_s:
        return None
    values = tuple("" if v is None else str(v) for v in raw)
    return ExcelRow(
        excel_row=excel_row,
        store=store_s,
        name=name_s,
        phone=phone_s,
        role=role_s if role_s else None,
        values=values,
    )


def parse_rows(
    rows,
    *,
    has_header: bool = True,
    col_store: int = 0,
    col_name: int = 1,
    col_phone: int = 3,
    col_role: int = 2,
    require_store: bool = False,
    require_name: bool = True,
    require_phone: bool = True,
):
    """批量解析 openpyxl 行（iter_rows(values_only=True) 的结果）。

    返回 (valid: list[ExcelRow], skipped: list[SkippedRow])：
      valid  —— 通过校验的行；
      skipped—— 被跳过的行（SkippedRow：物理行号 + 原因）。
    """
    if rows is None:
        rows = []
    start = 1 if has_header else 0
    valid = []
    skipped = []
    for offset, raw in enumerate(rows[start:]):
        physical = start + offset + 1  # 1 起物理行号（含表头行）
        r = parse_row(
            raw,
            excel_row=physical,
            col_store=col_store,
            col_name=col_name,
            col_phone=col_phone,
            col_role=col_role,
            require_store=require_store,
            require_name=require_name,
            require_phone=require_phone,
        )
        if r is None:
            reason = _skip_reason(raw, col_store=col_store, col_name=col_name,
                                  col_phone=col_phone, require_store=require_store,
                                  require_name=require_name, require_phone=require_phone)
            skipped.append(SkippedRow(excel_row=physical, reason=reason))
        else:
            valid.append(r)
    return valid, skipped


def _skip_reason(raw, *, col_store, col_name, col_phone,
                 require_store, require_name, require_phone) -> str:
    """生成跳过原因（用于 SkippedRow.reason）。"""
    n = len(raw) if raw is not None else 0

    def cell(idx):
        return raw[idx] if raw is not None and 0 <= idx < n else None

    store = cell(col_store)
    name = cell(col_name)
    phone = cell(col_phone)
    store_s = str(store).strip() if store is not None else ""
    name_s = str(name).strip() if name is not None else ""
    phone_s = format_phone(phone)
    if require_store and not store_s:
        if store is None:
            return "门店为空（门店用于 Logo 功能）"
        return f"门店为空（原值 {store!r} 为空白）"
    if require_name and not name_s:
        if name is None:
            return "姓名为空（请填写姓名）"
        return f"姓名为空（原值 {name!r} 为空白）"
    if require_phone and not phone_s:
        if phone is None:
            return "电话为空（若为公式单元格，请先在 Excel 中计算并保存）"
        return f"电话为空（原值 {phone!r}；若为公式单元格，请先在 Excel 中计算并保存）"
    return "行数据不完整"


def unique_stores(rows, col_store: int = 0):
    """从 openpyxl 行中提取去重门店名（保序，跳过空值）。"""
    out = []
    for r in rows:
        if r and len(r) > col_store and r[col_store] is not None:
            s = str(r[col_store]).strip()
            if s and s not in out:
                out.append(s)
    return out


# ---------------------------------------------------------------------------
# 统一入口
# ---------------------------------------------------------------------------
def load_excel_dataset(
    path: str,
    *,
    has_header: Optional[bool] = None,
    col_store: int = 0,
    col_name: int = 1,
    col_phone: int = 3,
    col_role: Optional[int] = 2,
    require_store: bool = True,
    require_name: bool = True,
    require_phone: bool = True,
) -> ExcelDataset:
    """统一 Excel 加载入口（GUI / CLI 全部走这里）。

    参数：
      path           : .xlsx / .xlsm 文件路径
      has_header     : None 表示「由调用方决定默认 True 并优先已有配置」；
                       True/False 显式指定。
      col_store      : 门店列（0 起）。默认 A=0。
      col_name       : 姓名列（0 起）。默认 B=1。
      col_phone      : 电话列（0 起）。默认 D=3。
      col_role       : 销售顾问列（0 起）；None / <0 表示不解析。
      require_store  : 门店是否必填（门店用于 Logo 功能；默认必填）
      require_name   : 姓名是否必填（默认必填）
      require_phone  : 电话是否必填（默认必填）

    返回 ExcelDataset；失败抛 ExcelDataError（区分文件不存在 / 格式不支持 /
    工作簿损坏 / 无工作表 / 列越界 / 无有效数据）。
    """
    if col_store == col_name:
        raise ExcelDataError("门店列与姓名列不能使用同一 Excel 列。", code="COL_CONFLICT")
    if col_name == col_phone:
        raise ExcelDataError("姓名和电话不能使用同一 Excel 列。", code="COL_CONFLICT")
    if col_store == col_phone:
        raise ExcelDataError("门店列与电话列不能使用同一 Excel 列。", code="COL_CONFLICT")

    if not path or not os.path.exists(path):
        raise ExcelDataError(f"Excel 文件不存在：{path}", code="FILE_NOT_FOUND")

    ext = os.path.splitext(path)[1].lower()
    if ext not in (".xlsx", ".xlsm"):
        raise ExcelDataError(
            f"不支持的文件格式 {ext or '(无扩展名)'}，请使用 .xlsx / .xlsm 文件（当前不支持 .xls）。",
            code="UNSUPPORTED_TYPE",
        )

    openpyxl = _get_openpyxl()
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    except Exception as e:
        raise ExcelDataError(f"无法打开 Excel 工作簿（可能已损坏）：{e}", code="WORKBOOK_BROKEN") from e

    try:
        if not wb.sheetnames:
            raise ExcelDataError("Excel 工作簿中没有工作表。", code="NO_SHEET")
        ws = wb.active
        sheet_name = ws.title
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    headers = []
    if has_header is None:
        has_header = True
    if has_header and rows:
        headers = ["" if v is None else str(v).strip() for v in rows[0]]

    max_columns = max((len(r) for r in rows), default=0)
    # 列索引越界检查（仅当工作簿非空时；空工作簿走 NO_VALID_DATA）
    if max_columns > 0:
        for label, idx in (("门店", col_store), ("姓名", col_name), ("电话", col_phone)):
            if idx >= max_columns and idx >= 0:
                raise ExcelDataError(
                    f"列 {index_to_excel_column(idx)}（{label}）超出工作表最大列数 {max_columns}（共 {max_columns} 列）。",
                    code="COL_OUT_OF_RANGE",
                )

    valid, skipped = parse_rows(
        rows,
        has_header=has_header,
        col_store=col_store,
        col_name=col_name,
        col_phone=col_phone,
        col_role=col_role,
        require_store=require_store,
        require_name=require_name,
        require_phone=require_phone,
    )

    # 门店列表：来自 dataset.stores（strip、保序、去重；门店用于 Logo 功能）
    stores = []
    for r in valid:
        s = r.store
        if s and s not in stores:
            stores.append(s)

    dataset = ExcelDataset(
        path=path,
        headers=headers,
        rows=rows,
        valid_rows=valid,
        skipped_rows=skipped,
        stores=stores,
        max_columns=max_columns,
        sheet_name=sheet_name,
    )

    if require_name or require_phone:
        if not valid:
            detail = ""
            if skipped:
                first = skipped[0]
                detail = f"（首条跳过原因：第 {first.excel_row} 行 {first.reason}）"
            raise ExcelDataError(
                f"Excel 中没有可生成的有效数据。{detail}", code="NO_VALID_DATA")
    return dataset


# ---------------------------------------------------------------------------
# 输出分组：Stage 4.5 已迁至 core/output_paths.py
# ---------------------------------------------------------------------------
# Excel 任意列分组输出（GroupFolderMap / resolve_output_directory /
# sanitize_group_component / OutputPathError）属于 output path 范畴，
# 统一放在 core/output_paths.py，避免本文件变成万能工具文件。
# ExcelRow.values 仍是分组值的唯一数据源（红线不变）。
