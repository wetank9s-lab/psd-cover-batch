#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
psd_cover_batch.py

用 Excel 表格里的数据，批量替换 PSD 模板中的文字图层（姓名 / 电话 / 销售顾问），
并按「门店」自动切换对应的门店 Logo 图层，最后导出成一张张 PNG 封面图。

适用于：抖音 / 视频号等渠道的「活动海报 / 视频封面」批量制作。
典型场景：一张 PSD 模板 + N 个门店销售顾问，每人一张带自己姓名、电话、门店 Logo 的封面。

依赖：
  - Windows + 已安装 Adobe Photoshop（通过 COM 自动化驱动，无需手动打开界面操作）
  - Python 3.8+ ，安装依赖： pip install -r requirements.txt

用法示例：
  # 1) 先看 PSD 里有哪些图层（确认图层名，方便修改下面的映射）
  python psd_cover_batch.py --psd 8081.psd --xlsx 门店销售0803销售顾问.xlsx --inspect

  # 2) 先合成第 1 行数据做测试
  python psd_cover_batch.py --psd 8081.psd --xlsx 门店销售0803销售顾问.xlsx --row 1 --out ./out

  # 3) 批量：每一行数据导出一张 PNG
  python psd_cover_batch.py --psd 8081.psd --xlsx 门店销售0803销售顾问.xlsx --out ./out
"""

import os
import sys
import time
import argparse
import win32com.client
import openpyxl

# Stage 0：引入 core 纯函数（行为等价），供测试复用
from core import util as core_util
# Stage 1：Photoshop 安全资源管理（Session 只关自己 open/duplicate 的文档）
from core.photoshop import PhotoshopSession, com_retry
# Stage 2：唯一 LayerRef（index_path 精确定位，取消「图层名 = 图层 ID」）
from core.layer_index import (
    collect_layer_index, resolve_layer, ref_from_config, LayerRef,
)

# ============================================================
# PSD 图层名映射（按你的模板修改这里）
# 变更图层名后，用 --inspect 确认即可。
# ============================================================
NAME_LAYER = "姓 名 拷贝"      # 姓名文字图层（注意原模板里名字可能带空格 / 拷贝后缀）
PHONE_LAYER = "电话 拷贝"       # 电话文字图层
TITLE_LAYER = "销售顾问 拷贝"   # 职位 / 头衔文字图层（如「销售顾问」）

# Excel 列索引（0 开始）：A=门店, B=姓名, C=职位, D=电话
STORE_COL, NAME_COL, TITLE_COL, PHONE_COL = 0, 1, 2, 3

# 输出的门店 Logo 图层判定：图层名与 Excel 门店名「模糊包含」匹配（互含即可，
# 例如 Excel「康乐」可匹配 PSD「康乐电器」）。品牌 Logo（名字含 logo）始终显示。


# ---------- COM 调用重试（解决 Photoshop 「应用程序正忙」）----------
# Stage 1：统一委托 core.photoshop.com_retry（只对可重试 com_error 重试，
# 参数错误/编程错误立即抛出；不再产生第三套 retry 实现）。
def com_call(func, *a, retries=15, **k):
    return com_retry(func, *a, retries=retries, delay=0.4, backoff=1.4, **k)


def collect_layers(layers, registry, depth=0, parent=""):
    """递归收集所有图层 name -> 对象，并记录是否为组。

    Stage 2 起：仅用于 inspect() 的组判断与展示；
    不再承担 text / logo 的 identity（run() 改走 LayerIndex + LayerRef）。
    """
    for layer in layers:
        registry[layer.Name] = layer
        try:
            if layer.Layers.Count > 0:
                registry["__group__:" + layer.Name] = True
                collect_layers(layer.Layers, registry, depth + 1, layer.Name)
        except Exception:
            pass


def set_text(registry, layer_name, text):
    """（兼容 wrapper，旧 name-only 路径）按 name 设置文字。

    Stage 2 起 run() 不再调用；保留仅作向后兼容 / 未被删除前的最小驻留。
    """
    if layer_name in registry:
        try:
            registry[layer_name].TextItem.Contents = text
            return True
        except Exception as e:
            print(f"  [警告] 设置文字层 {layer_name!r} 失败: {e}")
            return False
    print(f"  [警告] 未找到文字层 {layer_name!r}，已跳过")
    return False


def _resolve_text_ref(index, name):
    """把旧 name 配置解析为 LayerRef（Stage 2 精确定位）。

    返回 (status, ref | None)：status 见 core.layer_index（VALID/MIGRATED/AMBIGUOUS/MISSING）。
    """
    from core.layer_index import rebind_layer_reference
    return rebind_layer_reference(index, name)


def set_text_ref(index, doc, ref, text, label=""):
    """按 LayerRef 精确定位并写文字（替代按 name 的 set_text）。"""
    if ref is None:
        return False
    try:
        layer = resolve_layer(doc, ref)
        layer.TextItem.Contents = text
        return True
    except Exception as e:
        print(f"  [警告] 设置文字层 {label or ref.display_path!r} 失败: {e}")
        return False


def sanitize(name):
    """清洗为 Windows 可用文件名片段（由 core.util 提供统一实现）。"""
    return core_util.sanitize_filename(name)


def inspect(psd_path, xlsx_path):
    # COM 初始化/反初始化由 PhotoshopSession 的 __enter__/__exit__ 统一负责
    with PhotoshopSession() as ps:
            ps.app.DisplayDialogs = 2
            doc = ps.open_document(psd_path)
            print(f"文档尺寸: {int(doc.Width)} x {int(doc.Height)}")
            print("\n图层树:")
            registry = {}

            def walk(layers, depth=0):
                for layer in layers:
                    is_text = False
                    try:
                        _ = layer.TextItem
                        is_text = True
                    except Exception:
                        pass
                    is_group = False
                    try:
                        if layer.Layers.Count > 0:
                            is_group = True
                    except Exception:
                        pass
                    tag = "GRP" if is_group else ("TXT" if is_text else "LAY")
                    print("  " * depth + f"[{tag}] {layer.Name!r}")
                    if is_group:
                        walk(layer.Layers, depth + 1)

            walk(doc.Layers)

            # 库存门店名
            stores = set()
            if xlsx_path and os.path.exists(xlsx_path):
                wb = openpyxl.load_workbook(xlsx_path, data_only=True)
                for r in list(wb.active.iter_rows(values_only=True))[1:]:
                    if r and r[STORE_COL] is not None:
                        stores.add(str(r[STORE_COL]).strip())
            brand_logos = [l.Name for l in doc.Layers
                           if "logo" in l.Name.lower()
                           and l.Name.strip() not in stores
                           and "__group__:" + l.Name not in registry]
            brand_set = set(brand_logos)
            store_map = {}
            for l in doc.Layers:
                if "__group__:" + l.Name in registry or l.Name in brand_set:
                    continue
                for s in stores:
                    if core_util.fuzzy_contains(l.Name, s):
                        store_map.setdefault(s, []).append(l.Name)
                        break
            print("\n匹配到的门店 Logo 图层（与 Excel 门店名「模糊包含」匹配）:")
            if store_map:
                for s in sorted(store_map):
                    print(f"  {s!r:12s} -> {store_map[s]}")
            else:
                print("  (无，请检查门店名是否与图层名存在包含关系)")
            print("\n品牌 Logo 图层（名字含 logo，每张封面都显示）:")
            print("  " + (", ".join(brand_logos) if brand_logos else "(无)"))
            # with 退出：Session 只关闭自己打开的 doc


def run(psd_path, xlsx_path, out_dir, row=None, brand_logos=None):
    os.makedirs(out_dir, exist_ok=True)
    # COM 初始化/反初始化由 PhotoshopSession 的 __enter__/__exit__ 统一负责
    with PhotoshopSession() as ps:
            ps.app.DisplayDialogs = 2
            doc = ps.open_document(psd_path)
            ps.app.ActiveDocument = doc

            # Stage 2：用 LayerIndex 建立唯一图层身份（不再用 name 当 ID）
            index = collect_layer_index(doc)

            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            rows = list(wb.active.iter_rows(values_only=True))
            data = rows[1:]
            stores = set(str(r[STORE_COL]).strip() for r in data if r and r[STORE_COL])
            # 品牌 Logo：名字含 "logo"（且不正好等于门店名）的图层 —— 每张封面都显示，
            # 例如「七方logo」「七方logo 拷贝」。可用 --brand-logo 显式指定覆盖自动识别。
            auto_brand = [r for r in index.layers
                          if "logo" in r.name.lower()
                          and r.name.strip() not in stores
                          and not r.is_group]
            if brand_logos:
                brand_logos = [r for r in index.layers
                               if r.name in brand_logos and not r.is_group]
            else:
                brand_logos = auto_brand
            brand_ids = {r.id for r in brand_logos}

            # 门店 Logo：与 Excel 门店名「模糊包含」匹配（互含即可，不要求完全一致）。
            # 例如 Excel 的「康乐」可匹配 PSD 的「康乐电器」，「九兴」可匹配「九兴电器」。
            # 用 LayerRef 保存：每个门店 -> 匹配到的 PSD 图层 ref 列表
            store_map = {}      # Excel 门店名 -> 匹配到的 LayerRef 列表
            store_logos = []    # 去重后的所有门店 Logo LayerRef
            seen_logo_ids = set()
            for r in index.layers:
                if r.is_group:
                    continue
                if r.id in brand_ids:
                    continue
                for s in stores:
                    if core_util.fuzzy_contains(r.name, s):
                        store_map.setdefault(s, []).append(r)
                        if r.id not in seen_logo_ids:
                            seen_logo_ids.add(r.id)
                            store_logos.append(r)
                        break
            print(f"检测到门店 Logo 图层 {len(store_logos)} 个: "
                  f"{[r.display_path for r in store_logos]}")
            for s in sorted(store_map):
                print(f"    门店 {s!r:12s} -> {[r.display_path for r in store_map[s]]}")
            print(f"品牌 Logo 图层（始终显示）{len(brand_logos)} 个: "
                  f"{[r.display_path for r in brand_logos]}")

            # 文字层：Stage 2 按 LayerRef 精确定位（旧 name 常量 -> rebind 到唯一 ref；
            # 若同名歧义则回退 name 并在运行时按「精确 index_path」处理）
            text_refs = {}
            for key, legacy in (("姓名", NAME_LAYER), ("电话", PHONE_LAYER),
                                ("销售顾问", TITLE_LAYER)):
                status, ref = _resolve_text_ref(index, legacy)
                text_refs[key] = ref
                if status in ("AMBIGUOUS", "MISSING") or ref is None:
                    print(f"  [警告] 文字层 {legacy!r} 无法唯一解析（{status}），该字段将不替换")

            todo = [row - 1] if row else list(range(len(data)))
            exported = 0
            for idx in todo:
                if idx < 0 or idx >= len(data):
                    print(f"[跳过] 行号越界: {idx + 1}")
                    continue
                r = data[idx]
                if not r or r[NAME_COL] is None:
                    print(f"[跳过] 第 {idx + 1} 行数据不完整")
                    continue
                store = str(r[STORE_COL]).strip() if r[STORE_COL] is not None else ""
                name = str(r[NAME_COL]).strip()
                title = str(r[TITLE_COL]).strip() if r[TITLE_COL] is not None else ""
                phone = r[PHONE_COL]
                if isinstance(phone, float):
                    phone = int(phone)
                phone = str(phone).strip()

                print(f"\n[{idx + 1}/{len(data)}] 门店={store!r} 姓名={name!r} 电话={phone!r}")
                set_text_ref(index, doc, text_refs.get("姓名"), name, label=NAME_LAYER)
                set_text_ref(index, doc, text_refs.get("电话"), phone, label=PHONE_LAYER)
                set_text_ref(index, doc, text_refs.get("销售顾问"), title, label=TITLE_LAYER)

                target_refs = store_map.get(store, [])
                target_ids = {t.id for t in target_refs}
                for rr in store_logos:
                    try:
                        layer = resolve_layer(doc, rr)
                        layer.Visible = (rr.id in target_ids)
                    except Exception as e:
                        print(f"  [警告] 定位门店 Logo {rr.display_path!r} 失败: {e}")
                for rr in brand_logos:
                    try:
                        layer = resolve_layer(doc, rr)
                        layer.Visible = True
                    except Exception as e:
                        print(f"  [警告] 定位品牌 Logo {rr.display_path!r} 失败: {e}")
                shown = [t.display_path for t in target_refs] if target_refs else "（无匹配，门店 Logo 全部隐藏）"
                print(f"  门店 Logo: 显示 {store!r} -> {shown}；品牌 Logo: 全部显示")

                fname = f"{idx + 1:03d}_{sanitize(store)}_{sanitize(name)}.png"
                out_path = os.path.join(out_dir, fname)
                opts = win32com.client.Dispatch("Photoshop.PNGSaveOptions")
                opts.Interlaced = False
                com_call(doc.SaveAs, out_path, opts, True)  # asCopy=True，不改动原 PSD
                print(f"  已导出 -> {out_path}")
                exported += 1
            # with 退出：Session 只关闭自己打开的 doc（模板）
            print(f"\n完成：共导出 {exported} 张 PNG 到 {out_dir}")


def main():
    ap = argparse.ArgumentParser(description="PSD + Excel 批量生成封面 PNG")
    ap.add_argument("--psd", required=True, help="PSD 模板路径")
    ap.add_argument("--xlsx", required=True, help="Excel 数据路径")
    ap.add_argument("--out", default="./out", help="PNG 输出目录")
    ap.add_argument("--row", type=int, default=0, help="只合成指定行（1 开始），0 表示全部")
    ap.add_argument("--brand-logo", action="append", default=[],
                    help="品牌 Logo 图层名（每张封面都显示，如 七方logo）。可重复指定；不指定则自动识别名字含 logo 的图层")
    ap.add_argument("--inspect", action="store_true", help="只打印图层树后退出")
    args = ap.parse_args()

    if args.inspect:
        inspect(args.psd, args.xlsx)
        return
    run(args.psd, args.xlsx, args.out, row=args.row or None, brand_logos=args.brand_logo or None)


if __name__ == "__main__":
    import win32com.client  # 延迟导入，便于在非 Windows 上给出友好报错
    main()
