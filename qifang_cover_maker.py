# -*- coding: utf-8 -*-
"""
七方视频封面批量制作
===================
通过 Adobe Photoshop (COM 自动化) 批量替换 PSD 中的文字图层（姓名 / 电话 / 销售顾问）
并按门店切换对应的 Logo 图层，导出视频封面图。

依赖：
  - 本机已安装 Adobe Photoshop (2020+)
  - Python 包：openpyxl, pywin32, tkinter(内置)
"""
import os
import sys
import json
import time
import threading

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

import win32com.client
import pythoncom
import openpyxl

# Stage 0：引入 core 纯函数（不依赖 COM / Tk），保持行为一致，供测试与复用
from core import util as core_util
# Stage 1：Photoshop 安全资源管理（Session 只关自己 open/duplicate 的文档，绝不关用户文档）
from core.photoshop import PhotoshopSession, PhotoshopSessionError
# Stage 2：唯一 LayerRef（index_path 精确定位，取消「图层名 = 图层 ID」）
from core.layer_index import (
    LayerRef, LayerIndex, collect_layer_index, resolve_layer,
    rebind_layer_reference, ref_from_config, serialize_ref,
    VALID, MIGRATED, AMBIGUOUS, MISSING,
)

# ---------------- Photoshop 资源管理 ----------------
# Stage 1 起由 core.photoshop.PhotoshopSession 统一管理：
#   - Session 实例持有 app_started_by_tool / owned_documents / initial_documents；
#   - 只关闭 owned documents（本工具 Open / Duplicate 出来的）；
#   - 绝不遍历 Photoshop 当前全部 Documents 执行 Close；
#   - 不再使用模块级粘性全局 _PS_LAUNCHED_BY_US；
#   - 第一版保守 Quit 策略：默认完全不调用 app.Quit()（用户数据安全优先）。


APP_TITLE = "七方视频封面批量制作"
CONFIG_NAME = "qifang_cover_config.json"

# 导出格式
FMT_PNG = "PNG"
FMT_JPG = "JPG"
FMT_PSD = "PSD"

# 文本图层固定名称（PSD 内）
LAYER_NAME = "姓名"
LAYER_PHONE = "电话"
LAYER_ROLE = "销售顾问"


# ----------------------------------------------------------------------------
# Photoshop 辅助函数
# ----------------------------------------------------------------------------
def collect_layer_names(doc):
    """递归收集文档中所有图层。返回 [(name, is_group, parent_name), ...]（去重保序）。"""
    out = []

    def walk(parent, parent_name=""):
        try:
            n = parent.Layers.Count
        except Exception:
            return
        for i in range(1, n + 1):
            try:
                L = parent.Layers[i]
            except Exception:
                continue
            try:
                is_group = L.Layers.Count > 0
            except Exception:
                is_group = False
            out.append((L.Name, bool(is_group), parent_name))
            if is_group:
                walk(L, L.Name)

    walk(doc)
    # 去重但保持顺序
    seen = set()
    res = []
    for t in out:
        if t[0] not in seen:
            seen.add(t[0])
            res.append(t)
    return res


def find_layer(doc, target):
    """递归按名称查找图层对象，忽略名称内的空格差异。返回 Layer 或 None。"""
    t = target.replace(" ", "")

    def walk(parent):
        try:
            n = parent.Layers.Count
        except Exception:
            return None
        for i in range(1, n + 1):
            try:
                L = parent.Layers[i]
            except Exception:
                continue
            if L.Name.replace(" ", "") == t:
                return L
            try:
                if L.Layers.Count > 0:
                    r = walk(L)
                    if r is not None:
                        return r
            except Exception:
                pass
        return None

    return walk(doc)


def find_text_layer(doc, target):
    """查找文字图层，找不到或不是文字层时抛出清晰的中文错误。"""
    L = find_layer(doc, target)
    if L is None:
        raise RuntimeError(f"在 PSD 中找不到文字图层「{target}」（已尝试忽略空格差异）。")
    try:
        _ = L.TextItem.Contents
    except Exception:
        raise RuntimeError(f"图层「{target}」存在，但它不是文字图层，无法写入文本。")
    return L


def collect_text_layer_names(doc):
    """递归收集所有文字图层名称（去重保序）。"""
    out = []

    def walk(parent):
        try:
            n = parent.Layers.Count
        except Exception:
            return
        for i in range(1, n + 1):
            try:
                L = parent.Layers[i]
            except Exception:
                continue
            try:
                _ = L.TextItem.Contents
                if L.Name not in out:
                    out.append(L.Name)
            except Exception:
                pass
            try:
                if L.Layers.Count > 0:
                    walk(L)
            except Exception:
                pass

    walk(doc)
    return out


def safe_replace_text(doc, name, value, log=None):
    """安全替换文字图层文本：找不到或非文字层时记录警告并跳过，不抛异常。"""
    if not name:
        return False
    L = find_layer(doc, name)
    if L is None:
        if log:
            log(f"  警告：找不到文字图层「{name}」，已跳过该字段。")
        return False
    try:
        _ = L.TextItem.Contents
    except Exception:
        if log:
            log(f"  警告：图层「{name}」不是文字图层，已跳过该字段。")
        return False
    set_text_safe(L, value)
    return True


def set_layer_visible_by_name(doc, target, visible):
    """递归按名称设置图层可见性（只匹配第一个同名图层）。
    设为可见时，一并启用其所有父组，确保该图层真正显示出来。"""
    found = [False]

    def walk(parent):
        try:
            n = parent.Layers.Count
        except Exception:
            return
        for i in range(1, n + 1):
            try:
                L = parent.Layers[i]
            except Exception:
                continue
            if not found[0] and L.Name == target:
                try:
                    L.Visible = visible
                    found[0] = True
                    if visible:
                        _enable_parents(parent)
                except Exception:
                    pass
                return
            try:
                if L.Layers.Count > 0:
                    walk(L)
            except Exception:
                pass

    walk(doc)


def _ref_key(x):
    """把配置项（LayerRef dict / 旧 name 字符串 / LayerRef）归一化为唯一 key（id 优先，退化 name）。"""
    if isinstance(x, LayerRef):
        return x.id if x.id else x.name
    r = ref_from_config(x)
    if r is None:
        return None
    return r.id if r.id else r.name


def brand_logos_for(logo_layers, store_logo_map):
    """品牌 Logo：名字含 'logo' 且未被用作任何门店 Logo 的图层 —— 每张封面强制显示。

    与门店 Logo 区分：门店 Logo 会被映射给某个门店（出现在 store_logo_map 的值里），
    按门店切换显隐；品牌 Logo（如「七方logo」）不属于任何门店，应始终可见。

    Stage 2 语义：门店映射判定用「name 级别」——只要该 name 有任一图层被映射为门店，
    所有同名图层都视为门店候选（避免同名 Logo 一个被映射、另一个被误判为品牌）。
    logo_layers / store_logo_map 可能是 LayerRef dict 或旧 name 字符串。
    """
    mapped_names = set()
    for v in store_logo_map.values():
        r = ref_from_config(v)
        if r is not None:
            mapped_names.add(r.name)
    out = []
    for ln in logo_layers:
        r = ref_from_config(ln)
        if r is None:
            continue
        nm = r.name or ""
        if nm.lower().find("logo") >= 0 and nm not in mapped_names:
            out.append(ln)
    return out


def _fuzzy_contains(a, b):
    """门店名/图层名互含匹配（归一化后互相包含）。由 core.util.fuzzy_contains 提供。"""
    return core_util.fuzzy_contains(a, b)


def _enable_parents(container):
    """启用 container 及其所有祖先组（直到文档层）。"""
    p = container
    while True:
        try:
            p.Visible = True
            p = p.Parent
        except Exception:
            break


# ----------------------------------------------------------------------------
# Stage 2：LayerRef 精确操作辅助（替代「按 name 找第一个」的运行时定位）
# ----------------------------------------------------------------------------
def _find_indexed_layer(doc, index, ref_or_name):
    """把配置值（LayerRef dict / 旧 name 字符串 / LayerRef 对象）解析为 COM Layer。

    - 优先按 LayerRef.index_path 精确定位（resolve_layer，绝不 fallback 同名）；
    - 纯 name（旧配置 / 无 index_path）时走 rebind：唯一 -> 用之；歧义/缺失 -> None；
    - 返回 (layer, status)；status 用于日志（VALID/MIGRATED/AMBIGUOUS/MISSING/STALE）。
    """
    from core.layer_index import resolve_layer
    if isinstance(ref_or_name, LayerRef):
        ref = ref_or_name
    else:
        ref = ref_from_config(ref_or_name)
    if ref is None:
        return None, MISSING
    if ref.index_path:
        try:
            return resolve_layer(doc, ref), VALID
        except Exception as e:
            # index_path 失效（结构变化）：尝试按 name rebind，标记 MIGRATED/AMBIGUOUS
            status, cand = rebind_layer_reference(index, ref.name)
            if cand is not None:
                try:
                    return resolve_layer(doc, cand), MIGRATED
                except Exception:
                    return None, status
            return None, status
    # 纯 name（旧配置）
    status, cand = rebind_layer_reference(index, ref.name)
    if cand is None:
        return None, status
    try:
        return resolve_layer(doc, cand), status
    except Exception:
        return None, status


def set_text_by_ref(doc, index, ref_or_name, value, log=None, label=""):
    """按 LayerRef 精确替换文字层（找不到/歧义时记日志并跳过，不抛异常）。"""
    if not ref_or_name:
        return False
    layer, status = _find_indexed_layer(doc, index, ref_or_name)
    if layer is None:
        if log:
            reason = {"AMBIGUOUS": "同名图层多个，无法唯一确定",
                      "MISSING": "找不到该图层",
                      "MIGRATED": "已按唯一名称迁移"}.get(status, str(status))
            log(f"  警告：文字层 {label or ref_or_name} 定位失败（{reason}），已跳过该字段。")
        return False
    try:
        _ = layer.TextItem.Contents
    except Exception:
        if log:
            log(f"  警告：图层 {label or ref_or_name} 不是文字图层，已跳过该字段。")
        return False
    return set_text_safe(layer, value)


def set_visible_by_ref(doc, index, ref_or_name, visible, log=None, label=""):
    """按 LayerRef 精确设置图层可见性；可见时启用父组。返回 (ok, status)。"""
    if ref_or_name is None:
        return False, MISSING
    layer, status = _find_indexed_layer(doc, index, ref_or_name)
    if layer is None:
        return False, status
    try:
        layer.Visible = visible
        if visible:
            try:
                _enable_parents(layer.Parent)
            except Exception:
                pass
        return True, status
    except Exception as e:
        if log:
            log(f"  警告：设置可见性失败：{label or ref_or_name}：{e}")
        return False, status


def set_text_safe(layer, text, retries=6):
    for attempt in range(retries):
        try:
            layer.TextItem.Contents = text
            return True
        except Exception:
            if attempt == retries - 1:
                return False
            time.sleep(0.3 * (attempt + 1))


def export_doc(doc, path, fmt):
    if fmt == FMT_PNG:
        opt = win32com.client.Dispatch("Photoshop.PNGSaveOptions")
        opt.Interlaced = False
        opt.Compression = 6
    elif fmt == FMT_JPG:
        opt = win32com.client.Dispatch("Photoshop.JPGSaveOptions")
        opt.Quality = 9
        opt.EmbedColorProfile = False
    else:  # PSD
        opt = win32com.client.Dispatch("Photoshop.PhotoshopSaveOptions")
    doc.SaveAs(path, opt, True)


# ----------------------------------------------------------------------------
# 与 GUI 解耦的核心批处理逻辑（在 worker 线程调用）
# ----------------------------------------------------------------------------
def run_batch(cfg, progress_cb, log_cb, stop_flag):
    """
    cfg 字段：
      psd_path, xlsx_path, out_dir, has_header(bool),
      col_store, col_name, col_phone, col_role (0-based; col_role=-1 表示不替换),
      text_map (dict: "姓名"/"电话"/"销售顾问" -> 实际图层名 或 ""表示不替换),
      logo_layers (list[str]), store_logo_map (dict store->logo_layer or ""),
      fmt, also_png(bool)
    """
    # COM 初始化/反初始化由 PhotoshopSession 的 __enter__/__exit__ 统一负责
    try:
        # ---- 读取 Excel ----
        wb = openpyxl.load_workbook(cfg["xlsx_path"], data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        start = 1 if cfg["has_header"] else 0
        data = []
        skipped = 0
        for r in rows[start:]:
            store = r[cfg["col_store"]] if cfg["col_store"] < len(r) else None
            name = r[cfg["col_name"]] if cfg["col_name"] < len(r) else None
            phone = r[cfg["col_phone"]] if cfg["col_phone"] < len(r) else None
            if name is None or phone is None or str(name).strip() == "":
                skipped += 1
                continue
            data.append(r)
        log_cb(f"Excel 读取完成：{len(data)} 行有效数据，跳过 {skipped} 行空行。")
        if not data:
            log_cb("没有可处理的数据，已停止。")
            return

        # ---- 连接 Photoshop（Session 只管理自己打开/复制的文档）----
        log_cb("正在启动 / 连接 Photoshop ...")
        with PhotoshopSession() as ps:
            app = ps.app
            doc0 = ps.open_document(cfg["psd_path"])
            time.sleep(0.6)

            # Stage 2：运行时建立 LayerIndex（对当前打开的模板文档）
            index = collect_layer_index(doc0)

            text_map = cfg.get("text_map", {})
            name_ref = text_map.get("姓名", "")
            phone_ref = text_map.get("电话", "")
            role_ref = text_map.get("销售顾问", "")

            log_cb(f"模板已加载（{doc0.Width}x{doc0.Height}）。"
                   f"文字层映射：姓名→{name_ref.get('display_path') if isinstance(name_ref, dict) else (name_ref or '（不替换）')}，"
                   f"电话→{phone_ref.get('display_path') if isinstance(phone_ref, dict) else (phone_ref or '（不替换）')}，"
                   f"销售顾问→{role_ref.get('display_path') if isinstance(role_ref, dict) else (role_ref or '（不替换）')}。")

            out_dir = cfg["out_dir"]
            os.makedirs(out_dir, exist_ok=True)
            logo_layers = cfg["logo_layers"]
            store_logo_map = cfg["store_logo_map"]
            fmt = cfg["fmt"]
            also_png = cfg["also_png"]
            col_role = cfg["col_role"]
            col_name = cfg["col_name"]
            col_phone = cfg["col_phone"]
            col_store = cfg["col_store"]

            total = len(data)
            done = 0
            t0 = time.time()
            for idx, r in enumerate(data, start=1):
                if stop_flag.is_set():
                    log_cb("用户已停止。")
                    break
                store = str(r[col_store]).strip() if r[col_store] is not None else ""
                name = str(r[col_name]).strip()
                phone_v = r[col_phone]
                phone_s = str(int(phone_v)) if isinstance(phone_v, (int, float)) else str(phone_v).strip()
                if col_role >= 0 and col_role < len(r) and r[col_role] is not None:
                    role = str(r[col_role]).strip()
                else:
                    role = None

                d = None
                try:
                    d = ps.duplicate_document(doc0)
                    app.ActiveDocument = d
                    time.sleep(0.05)

                    # Stage 2：按 LayerRef 精确定位（同 LayerRef 在 duplicate 上重新 resolve，
                    # 绝不把 template 的 COM layer 拿去操作 duplicate）
                    if name_ref:
                        set_text_by_ref(d, index, name_ref, name, log_cb, label="姓名")
                    # 电话
                    if phone_ref:
                        set_text_by_ref(d, index, phone_ref, phone_s, log_cb, label="电话")
                    # 销售顾问（仅当 Excel 该列有值）
                    if role_ref and role is not None:
                        set_text_by_ref(d, index, role_ref, role, log_cb, label="销售顾问")

                    # Logo 图层切换：只显示本门店对应的 Logo，隐藏其余候选 Logo；
                    # 品牌 Logo（含 'logo' 且非门店 Logo）每张封面强制显示。
                    # Stage 2：identity 已改为 LayerRef（layer_id），业务规则（品牌/门店）不变。
                    mapped = store_logo_map.get(store, "")
                    brand = brand_logos_for(cfg["logo_layers"], cfg["store_logo_map"])
                    brand_keys = {_ref_key(x) for x in brand}
                    mapped_key = _ref_key(mapped) if mapped else None
                    for ln in cfg["logo_layers"]:
                        if _ref_key(ln) in brand_keys:
                            set_visible_by_ref(d, index, ln, True, log_cb, label="品牌Logo")
                        else:
                            set_visible_by_ref(d, index, ln, (_ref_key(ln) == mapped_key),
                                               log_cb, label="门店Logo")

                    safe_store = core_util.sanitize_filename(store)
                    safe_name = core_util.sanitize_filename(name)
                    base = f"{idx:03d}_{safe_store}_{safe_name}"
                    if fmt == FMT_PSD:
                        p = os.path.join(out_dir, base + ".psd")
                        export_doc(d, p, FMT_PSD)
                        if also_png:
                            export_doc(d, os.path.join(out_dir, base + ".png"), FMT_PNG)
                    else:
                        p = os.path.join(out_dir, base + "." + fmt.lower())
                        export_doc(d, p, fmt)

                    done += 1
                    if done % 5 == 0 or done == total:
                        el = time.time() - t0
                        progress_cb(done, total)
                        log_cb(f"  已生成 {done}/{total}  （{el:.1f}s）")
                finally:
                    # 无论本行成功/失败，都关闭本工具创建的 duplicate，绝不波及用户文档
                    if d is not None:
                        try:
                            ps.close_owned_document(d)
                        except Exception as e:
                            log_cb(f"  警告：关闭副本失败：{e}")

            el = time.time() - t0
            log_cb(f"完成！共生成 {done} 张，耗时 {el:.1f}s。输出目录：{out_dir}")
    except Exception as e:
        log_cb(f"错误：{e}")
        import traceback
        log_cb(traceback.format_exc())


# ----------------------------------------------------------------------------
# GUI
# ----------------------------------------------------------------------------
class App:
    def __init__(self, root):
        self.root = root
        root.title(APP_TITLE)
        root.geometry("820x760")
        root.resizable(True, True)

        self.config_path = self._find_config_path()
        self.cfg = {
            "psd_path": "", "xlsx_path": "", "out_dir": "",
            "has_header": False,
            "col_store": 0, "col_name": 1, "col_phone": 3, "col_role": 2,
            "text_map": {"姓名": "", "电话": "", "销售顾问": ""},
            "logo_layers": [], "store_logo_map": {},
            "fmt": FMT_PNG, "also_png": False,
        }
        # 运行时状态
        self.all_psd_layers = []
        self.all_psd_is_group = {}
        self.all_psd_parent = {}
        self.all_text_layers = []
        # Stage 2：LayerRef 状态（唯一图层身份）
        self.layer_index = None          # 当前 PSD 的 LayerIndex（不丢同名）
        self._ref_by_id = {}             # layer_id -> LayerRef
        self.layer_labels = {}           # layer_id -> 展示 label（同名时含 [id=...] 后缀，保证肉眼可区分）
        self.text_label_to_ref = {}      # label -> LayerRef（文字层）
        self.logo_label_to_ref = {}      # label -> LayerRef（当前有效 Logo 叶子）
        self.excel_stores = []
        self.excel_headers = []
        self.running = False
        self.stop_flag = threading.Event()
        self.worker = None

        self._build_ui()
        self._load_config()

    # ---------------- 路径 ----------------
    def _find_config_path(self):
        if getattr(sys, "frozen", False):
            base = os.path.dirname(sys.executable)
        else:
            base = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(base, CONFIG_NAME)

    # ---------------- UI 构建 ----------------
    def _build_ui(self):
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=8, pady=8)

        # Tab 1: 文件与列映射
        f1 = ttk.Frame(self.notebook)
        self.notebook.add(f1, text="① 文件与字段")
        self._build_tab_files(f1)

        # Tab 2: Logo 映射
        f2 = ttk.Frame(self.notebook)
        self.notebook.add(f2, text="② 门店Logo映射")
        self._build_tab_logo(f2)

        # Tab 3: 运行
        f3 = ttk.Frame(self.notebook)
        self.notebook.add(f3, text="③ 生成")
        self._build_tab_run(f3)

    def _build_tab_files(self, parent):
        # 文件选择
        row = ttk.LabelFrame(parent, text="文件选择", padding=8)
        row.pack(fill="x", padx=6, pady=6)

        ttk.Label(row, text="PSD 模板:").grid(row=0, column=0, sticky="e", pady=3)
        self.psd_var = tk.StringVar()
        ttk.Entry(row, textvariable=self.psd_var, width=55).grid(row=0, column=1, padx=4)
        ttk.Button(row, text="浏览...", command=self._pick_psd).grid(row=0, column=2)

        ttk.Label(row, text="Excel 数据:").grid(row=1, column=0, sticky="e", pady=3)
        self.xlsx_var = tk.StringVar()
        ttk.Entry(row, textvariable=self.xlsx_var, width=55).grid(row=1, column=1, padx=4)
        ttk.Button(row, text="浏览...", command=self._pick_xlsx).grid(row=1, column=2)

        ttk.Label(row, text="输出目录:").grid(row=2, column=0, sticky="e", pady=3)
        self.out_var = tk.StringVar()
        ttk.Entry(row, textvariable=self.out_var, width=55).grid(row=2, column=1, padx=4)
        ttk.Button(row, text="浏览...", command=self._pick_out).grid(row=2, column=2)

        self.header_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(row, text="Excel 首行为表头（数据从第2行开始）",
                        variable=self.header_var).grid(row=3, column=1, sticky="w", pady=3)

        # 字段映射
        mf = ttk.LabelFrame(parent, text="字段映射（选择 Excel 列）", padding=8)
        mf.pack(fill="x", padx=6, pady=6)

        cols = [chr(ord("A") + i) for i in range(16)]  # A..P
        self.col_store_var = tk.StringVar(value="A")
        self.col_name_var = tk.StringVar(value="B")
        self.col_phone_var = tk.StringVar(value="D")
        self.col_role_var = tk.StringVar(value="C")

        ttk.Label(mf, text="门店列（用于选 Logo）:").grid(row=0, column=0, sticky="e", pady=3)
        ttk.Combobox(mf, textvariable=self.col_store_var, values=cols, width=6,
                     state="readonly").grid(row=0, column=1, padx=4, sticky="w")
        ttk.Label(mf, text="姓名列:").grid(row=0, column=2, sticky="e", pady=3)
        ttk.Combobox(mf, textvariable=self.col_name_var, values=cols, width=6,
                     state="readonly").grid(row=0, column=3, padx=4, sticky="w")
        ttk.Label(mf, text="电话列:").grid(row=1, column=0, sticky="e", pady=3)
        ttk.Combobox(mf, textvariable=self.col_phone_var, values=cols, width=6,
                     state="readonly").grid(row=1, column=1, padx=4, sticky="w")
        ttk.Label(mf, text="销售顾问列:").grid(row=1, column=2, sticky="e", pady=3)
        ttk.Combobox(mf, textvariable=self.col_role_var, values=cols + ["（不替换）"], width=10,
                     state="readonly").grid(row=1, column=3, padx=4, sticky="w")

        ttk.Label(mf, text="提示：销售顾问列选「不替换」则保留 PSD 原文字。").grid(
            row=2, column=0, columnspan=4, sticky="w", pady=(4, 0))

        # 文字图层映射（PSD 内实际文字图层 -> 字段）
        tf = ttk.LabelFrame(parent, text="文字图层映射（PSD 内实际文字图层，加载后自动识别）", padding=8)
        tf.pack(fill="x", padx=6, pady=6)

        self.tm_name_var = tk.StringVar(value="（不替换）")
        self.tm_phone_var = tk.StringVar(value="（不替换）")
        self.tm_role_var = tk.StringVar(value="（不替换）")
        tm_opts = ["（不替换）"]

        ttk.Label(tf, text="姓名 →").grid(row=0, column=0, sticky="e", pady=3)
        self.tm_name_cb = ttk.Combobox(tf, textvariable=self.tm_name_var, values=tm_opts,
                                       width=30, state="readonly")
        self.tm_name_cb.grid(row=0, column=1, padx=4, sticky="w")
        ttk.Label(tf, text="电话 →").grid(row=0, column=2, sticky="e", pady=3)
        self.tm_phone_cb = ttk.Combobox(tf, textvariable=self.tm_phone_var, values=tm_opts,
                                        width=30, state="readonly")
        self.tm_phone_cb.grid(row=0, column=3, padx=4, sticky="w")
        ttk.Label(tf, text="销售顾问 →").grid(row=1, column=0, sticky="e", pady=3)
        self.tm_role_cb = ttk.Combobox(tf, textvariable=self.tm_role_var, values=tm_opts,
                                       width=30, state="readonly")
        self.tm_role_cb.grid(row=1, column=1, padx=4, sticky="w")

        ttk.Label(tf, text="提示：若 PSD 内文字层名与「姓名/电话/销售顾问」不同（如带空格），"
                            "在此手动指定；没有对应图层请保持「不替换」。").grid(
            row=2, column=0, columnspan=4, sticky="w", pady=(4, 0))

        # 加载按钮
        bf = ttk.Frame(parent)
        bf.pack(fill="x", padx=6, pady=6)
        ttk.Button(bf, text="加载 PSD / Excel 并分析图层", command=self._load).pack(side="left")
        ttk.Button(bf, text="保存配置", command=self._save_config).pack(side="left", padx=6)

    def _build_tab_logo(self, parent):
        info = ttk.Label(parent, text="先到「① 文件与字段」点击【加载】按钮，解析出 PSD 图层与 Excel 门店后再配置。")
        info.pack(fill="x", padx=8, pady=8)
        self.logo_info = info

        # 左右分栏
        paned = ttk.PanedWindow(parent, orient="horizontal")
        paned.pack(fill="both", expand=True, padx=6, pady=6)

        left = ttk.Frame(paned)
        right = ttk.Frame(paned)
        paned.add(left, weight=1)
        paned.add(right, weight=1)

        # 左：Logo 图层清单（可勾选）
        lf = ttk.LabelFrame(left, text="标记为 Logo 图层（PSD 内的候选 Logo）", padding=6)
        lf.pack(fill="both", expand=True, padx=4, pady=4)
        self.logo_canvas = tk.Canvas(lf)
        self.logo_scroll = ttk.Scrollbar(lf, orient="vertical", command=self.logo_canvas.yview)
        self.logo_inner = ttk.Frame(self.logo_canvas)
        self.logo_inner.bind("<Configure>", lambda e: self.logo_canvas.configure(
            scrollregion=self.logo_canvas.bbox("all")))
        self.logo_canvas.create_window((0, 0), window=self.logo_inner, anchor="nw")
        self.logo_canvas.configure(yscrollcommand=self.logo_scroll.set)
        self.logo_canvas.pack(side="left", fill="both", expand=True)
        self.logo_scroll.pack(side="right", fill="y")
        self.logo_checks = {}  # name -> BooleanVar

        # 右：门店 -> Logo 映射
        rf = ttk.LabelFrame(right, text="门店 → Logo 图层 映射", padding=6)
        rf.pack(fill="both", expand=True, padx=4, pady=4)
        self.map_canvas = tk.Canvas(rf)
        self.map_scroll = ttk.Scrollbar(rf, orient="vertical", command=self.map_canvas.yview)
        self.map_inner = ttk.Frame(self.map_canvas)
        self.map_inner.bind("<Configure>", lambda e: self.map_canvas.configure(
            scrollregion=self.map_canvas.bbox("all")))
        self.map_canvas.create_window((0, 0), window=self.map_inner, anchor="nw")
        self.map_canvas.configure(yscrollcommand=self.map_scroll.set)
        self.map_canvas.pack(side="left", fill="both", expand=True)
        self.map_scroll.pack(side="right", fill="y")
        self.map_combos = {}  # store -> StringVar
        self.map_combo_widgets = {}  # store -> Combobox widget

    def _build_tab_run(self, parent):
        # 选项
        opt = ttk.LabelFrame(parent, text="导出选项", padding=8)
        opt.pack(fill="x", padx=6, pady=6)
        self.fmt_var = tk.StringVar(value=FMT_PNG)
        ttk.Label(opt, text="导出格式:").grid(row=0, column=0, sticky="e", pady=3)
        ttk.Combobox(opt, textvariable=self.fmt_var, values=[FMT_PNG, FMT_JPG, FMT_PSD],
                     width=8, state="readonly").grid(row=0, column=1, sticky="w", padx=4)
        self.also_png_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(opt, text="导出 PSD 时另外再生成 PNG（便于预览）",
                        variable=self.also_png_var).grid(row=0, column=2, sticky="w", padx=10)

        # 进度
        pf = ttk.Frame(parent)
        pf.pack(fill="x", padx=6, pady=4)
        self.progress = ttk.Progressbar(pf, mode="determinate", maximum=100)
        self.progress.pack(fill="x", side="left", expand=True)
        self.progress_label = ttk.Label(pf, text="0 / 0")
        self.progress_label.pack(side="left", padx=6)

        # 按钮
        bf = ttk.Frame(parent)
        bf.pack(fill="x", padx=6, pady=4)
        self.btn_run = ttk.Button(bf, text="▶ 开始生成", command=self._start)
        self.btn_run.pack(side="left")
        self.btn_stop = ttk.Button(bf, text="■ 停止", command=self._stop, state="disabled")
        self.btn_stop.pack(side="left", padx=6)
        self.btn_preview = ttk.Button(bf, text="🖼 试做第1张预览", command=self._preview)
        self.btn_preview.pack(side="left", padx=6)

        # 日志
        lf = ttk.LabelFrame(parent, text="运行日志", padding=6)
        lf.pack(fill="both", expand=True, padx=6, pady=6)
        self.log = scrolledtext.ScrolledText(lf, height=12, state="disabled",
                                             font=("Consolas", 9))
        self.log.pack(fill="both", expand=True)

    # ---------------- 文件选择 ----------------
    def _pick_psd(self):
        p = filedialog.askopenfilename(title="选择 PSD 模板",
                                       filetypes=[("PSD", "*.psd"), ("所有", "*.*")])
        if p:
            self.psd_var.set(p)

    def _pick_xlsx(self):
        p = filedialog.askopenfilename(title="选择 Excel 数据",
                                       filetypes=[("Excel", "*.xlsx *.xls"), ("所有", "*.*")])
        if p:
            self.xlsx_var.set(p)
            # 默认输出目录
            if not self.out_var.get():
                d = os.path.splitext(os.path.basename(p))[0] + "_封面输出"
                self.out_var.set(os.path.join(os.path.dirname(p), d))

    def _pick_out(self):
        d = filedialog.askdirectory(title="选择输出目录")
        if d:
            self.out_var.set(d)

    # ---------------- 加载解析 ----------------
    def _load(self):
        psd = self.psd_var.get().strip()
        xlsx = self.xlsx_var.get().strip()
        if not psd or not os.path.exists(psd):
            messagebox.showerror("错误", "请先选择有效的 PSD 模板文件。")
            return
        if not xlsx or not os.path.exists(xlsx):
            messagebox.showerror("错误", "请先选择有效的 Excel 数据文件。")
            return

        # 读取 Excel 门店 + 列头
        try:
            wb = openpyxl.load_workbook(xlsx, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            has_header = self.header_var.get()
            start = 1 if has_header else 0
            self.excel_headers = [chr(ord("A") + i) for i in range(max((len(r) for r in rows), default=0))]
            stores = []
            for r in rows[start:]:
                if r and r[0] is not None and str(r[0]).strip():
                    s = str(r[0]).strip()
                    if s not in stores:
                        stores.append(s)
            self.excel_stores = stores
        except Exception as e:
            messagebox.showerror("Excel 错误", str(e))
            return

        # 读取 PSD 图层（通过 Photoshop COM；Session 负责 COM 初始化并只关闭自己打开的文档）
        try:
            self._log_gui(f"正在用 Photoshop 解析 PSD 图层：{os.path.basename(psd)}")
            with PhotoshopSession() as ps:
                doc = ps.open_document(psd)
                time.sleep(0.6)
                # Stage 2：建立唯一 LayerIndex（不丢同名图层），替代旧的按 name 去重
                self.layer_index = collect_layer_index(doc)
                self._ref_by_id = {r.id: r for r in self.layer_index.layers}
                self.layer_labels = self.layer_index.labels()   # id -> 唯一 label
                # 兼容缓存（旧逻辑仍读取；新建逻辑优先用 layer_index）
                self.all_psd_layers = [r.name for r in self.layer_index.layers]
                self.all_psd_is_group = {r.id: r.is_group for r in self.layer_index.layers}
                self.all_psd_parent = {r.id: self._parent_name_of(r) for r in self.layer_index.layers}
                self.all_text_layers = [r.display_path for r in self.layer_index.layers if r.is_text]
                # 文字层 label -> LayerRef（同名 display_path 已由 labels 附加 id 后缀）
                self.text_label_to_ref = {}
                for r in self.layer_index.layers:
                    if r.is_text:
                        self.text_label_to_ref[self.layer_labels[r.id]] = r
                self.logo_label_to_ref = {}
                # with 退出：Session 只关闭自己打开的 doc（即上面的模板文档）
        except Exception as e:
            messagebox.showerror("PSD 错误", f"无法解析 PSD（请确认 Photoshop 已安装并可启动）：\n{e}")
            return

        # 默认勾选 Logo 候选：名称含 "logo"，或位于名为 logo 的组内，
        # 或图层名与 Excel 门店名一致（平铺式门店 Logo），且本身不是组
        def is_logo_candidate(ref):
            if ref.is_group:
                return False
            if "logo" in ref.name.lower():
                return True
            parent = (self._parent_name_of(ref) or "").lower()
            if "logo" in parent:
                return True
            return ref.name in stores

        # 恢复上次保存的 Logo 勾选状态（含勾选的组），否则按名称启发式默认
        # Stage 2：logo_layers 可能是 LayerRef dict 或旧 name 字符串；
        # dict 不可 hash，不能直接 set()，统一转成 name/id 判断。
        saved_items = self.cfg.get("logo_layers", [])
        saved_names = {x for x in saved_items if isinstance(x, str)}
        saved_logo_ids = {ref_from_config(x).id for x in saved_items
                          if ref_from_config(x) is not None and ref_from_config(x).id}
        saved_logo = saved_names
        children_map = self._build_children_map()

        def inherited_checked(ref):
            if ref.id in saved_logo_ids or ref.name in saved_names:
                return True
            parent = self._parent_name_of(ref)
            seen = set()
            while parent:
                if parent in saved_names or parent in saved_logo_ids:
                    return True
                if parent in seen:
                    break
                seen.add(parent)
                parent = self._parent_name_of_str(parent)
            return False

        self.logo_checks = {}
        for ref in self.layer_index.layers:
            label = self.layer_labels[ref.id]
            checked = is_logo_candidate(ref) or inherited_checked(ref)
            self.logo_checks[label] = tk.BooleanVar(value=checked)

        # 自动匹配 门店->Logo（图层名包含门店名则自动选；Stage 2：同名歧义不自动选第一个）
        prev_map = self.cfg.get("store_logo_map", {})
        self.map_combos = {}
        logo_candidate_refs = [r for r in self.layer_index.layers if is_logo_candidate(r)]
        for s in stores:
            default_label = "（无）"
            if s in prev_map:
                prev = prev_map[s]
                prev_ref = ref_from_config(prev)
                # 旧配置 name 唯一命中 -> 用；否则视为未配置
                if prev_ref is not None:
                    if prev_ref.id and prev_ref.id in self._ref_by_id:
                        default_label = self.layer_labels[prev_ref.id]
                    else:
                        m = self.layer_index.find_matching(prev_ref.name)
                        if len(m) == 1:
                            default_label = self.layer_labels[m[0].id]
                        elif len(m) > 1:
                            default_label = "（无）"   # ambiguous：不自动选
            else:
                hits = [r for r in logo_candidate_refs
                        if core_util.fuzzy_contains(r.name, s)]
                if len(hits) == 1:
                    default_label = self.layer_labels[hits[0].id]
                elif len(hits) > 1:
                    default_label = "（无）"   # 同名歧义：不自动选
            self.map_combos[s] = tk.StringVar(value=default_label)

        # 文字图层映射：Stage 2 候选显示 display_path（同名肉眼可区分）；
        # 自动匹配：名称（忽略空格）全局唯一才自动选，歧义保持「不替换」
        opts = ["（不替换）"] + [self.layer_labels[r.id] for r in self.layer_index.layers if r.is_text]
        self.tm_name_cb["values"] = opts
        self.tm_phone_cb["values"] = opts
        self.tm_role_cb["values"] = opts
        prev_tm = self.cfg.get("text_map", {})

        def auto_text(field):
            t = field.replace(" ", "")
            # 先用上次的显式配置（LayerRef dict 或旧 name）
            if prev_tm.get(field):
                prev = prev_tm[field]
                prev_ref = ref_from_config(prev)
                if prev_ref is not None:
                    if prev_ref.id and prev_ref.id in self._ref_by_id:
                        return self.layer_labels[prev_ref.id]
                    m = self.layer_index.find_matching(prev_ref.name)
                    if len(m) == 1:
                        return self.layer_labels[m[0].id]
                    return "（不替换）"   # ambiguous / missing
            # 再按名称（去空格）自动匹配：仅当全局唯一
            m = self.layer_index.find_matching(field)
            if len(m) == 1:
                return self.layer_labels[m[0].id]
            return "（不替换）"

        self.tm_name_var.set(auto_text("姓名"))
        self.tm_phone_var.set(auto_text("电话"))
        self.tm_role_var.set(auto_text("销售顾问"))

        self._rebuild_logo_lists()
        self._log_gui(f"解析完成：PSD 共 {len(self.layer_index)} 个图层，Excel 共 {len(stores)} 个门店。")
        self.logo_info.config(text=f"PSD 图层 {len(self.layer_index)} 个 ｜ Excel 门店 {len(stores)} 个 ｜ 请在下方勾选 Logo 并完成映射。")

    def _parent_name_of(self, ref):
        """返回 LayerRef 的父组名（display_path 倒数第二段）。"""
        if not ref or not ref.display_path:
            return ""
        parts = [p.strip() for p in ref.display_path.split(">")]
        return parts[-2] if len(parts) >= 2 else ""

    def _parent_name_of_str(self, name):
        """兼容：按 name 找 ref 再取父名（旧逻辑用）。"""
        if self.layer_index is None:
            return ""
        m = self.layer_index.find_matching(name)
        if len(m) == 1:
            return self._parent_name_of(m[0])
        return ""

    def _label_to_ref(self, label):
        """把展示 label（display_path 或 display_path [id=...]）解析回 LayerRef。"""
        if self.layer_index is None:
            return None
        if not label:
            return None
        # 优先精确 id 后缀
        for r in self.layer_index.layers:
            if self.layer_labels.get(r.id) == label:
                return r
        # 退化为 display_path 匹配
        for r in self.layer_index.layers:
            if r.display_path == label:
                return r
        return None

    def _build_children_map(self):
        """根据 all_psd_parent 构建 父名 -> [子名,...] 映射（兼容旧逻辑）。"""
        children = {}
        for rid, parent in self.all_psd_parent.items():
            if parent:
                children.setdefault(parent, []).append(rid)
        return children

    def _effective_logo_layers(self):
        """返回当前应作为候选 Logo 的『叶子图层 label 列表』（含被勾选组的子图层）。"""
        children = self._build_children_map()
        result = []

        def collect(rid):
            ref = self._ref_by_id.get(rid)
            if ref is None:
                return
            if ref.is_group:
                for ch in children.get(rid, []):
                    collect(ch)
            else:
                label = self.layer_labels.get(rid)
                if label and label not in result:
                    result.append(label)

        for label, var in self.logo_checks.items():
            if var.get():
                ref = self._label_to_ref(label)
                if ref is not None:
                    collect(ref.id)
        return result

    def _rebuild_logo_lists(self):
        # 清空
        for w in self.logo_inner.winfo_children():
            w.destroy()
        for w in self.map_inner.winfo_children():
            w.destroy()
        self.map_combo_widgets = {}

        ttk.Label(self.logo_inner, text="☑ 图层（显示完整路径）", font=("Microsoft YaHei", 9, "bold")).grid(
            row=0, column=0, sticky="w", padx=4, pady=2)
        for i, label in enumerate(self.logo_checks.keys(), start=1):
            cb = ttk.Checkbutton(self.logo_inner, text=label, variable=self.logo_checks[label])
            cb.grid(row=i, column=0, sticky="w", padx=4, pady=1)
            # 勾选状态变化时，同步刷新右侧门店→Logo 下拉列表
            self.logo_checks[label].trace_add(
                "write", lambda *a: self._on_logo_checks_changed())

        # 映射表
        ttk.Label(self.map_inner, text="门店", font=("Microsoft YaHei", 9, "bold")).grid(
            row=0, column=0, sticky="w", padx=4, pady=2)
        ttk.Label(self.map_inner, text="→ Logo 图层", font=("Microsoft YaHei", 9, "bold")).grid(
            row=0, column=1, sticky="w", padx=4, pady=2)
        logo_opts = ["（无）"] + self._effective_logo_layers()
        for i, s in enumerate(self.excel_stores, start=1):
            ttk.Label(self.map_inner, text=s).grid(row=i, column=0, sticky="w", padx=4, pady=1)
            cb = ttk.Combobox(self.map_inner, textvariable=self.map_combos[s],
                              values=logo_opts, width=40, state="readonly")
            cb.grid(row=i, column=1, sticky="w", padx=4, pady=1)
            self.map_combo_widgets[s] = cb

    def _on_logo_checks_changed(self):
        """左侧勾选 Logo 图层变动时，实时同步右侧门店→Logo 下拉的可选项。"""
        logo_opts = ["（无）"] + self._effective_logo_layers()
        for s, var in self.map_combos.items():
            w = self.map_combo_widgets.get(s)
            if w is None:
                continue
            cur = var.get()
            if cur not in logo_opts:
                var.set("（无）")
            w["values"] = logo_opts

    # ---------------- 运行 ----------------
    def _collect_cfg(self):
        col_of = lambda v: -1 if v == "（不替换）" else ord(v) - ord("A")
        # 保存用户显式勾选的图层（含组），用于下次加载时恢复
        # Stage 2：以 LayerRef dict 保存（layer_id/name/display_path），并保留 name 供旧版兼容
        def ref_to_cfg(label):
            ref = self._label_to_ref(label)
            if ref is None:
                return label   # 旧 name（极端情况）
            return serialize_ref(ref)

        logo_layers = [ref_to_cfg(label)
                       for label in self.all_psd_layers_labels()
                       if self.logo_checks.get(label, tk.BooleanVar()).get()]
        store_logo_map = {}
        for s, var in self.map_combos.items():
            val = var.get()
            store_logo_map[s] = "" if val == "（无）" else ref_to_cfg(val)
        def tm_val(var):
            v = var.get()
            if v == "（不替换）":
                return ""
            return ref_to_cfg(v)
        text_map = {
            "姓名": tm_val(self.tm_name_var),
            "电话": tm_val(self.tm_phone_var),
            "销售顾问": tm_val(self.tm_role_var),
        }
        return {
            "psd_path": self.psd_var.get().strip(),
            "xlsx_path": self.xlsx_var.get().strip(),
            "out_dir": self.out_var.get().strip(),
            "has_header": self.header_var.get(),
            "col_store": col_of(self.col_store_var.get()),
            "col_name": col_of(self.col_name_var.get()),
            "col_phone": col_of(self.col_phone_var.get()),
            "col_role": col_of(self.col_role_var.get()),
            "text_map": text_map,
            "logo_layers": logo_layers,
            "store_logo_map": store_logo_map,
            "fmt": self.fmt_var.get(),
            "also_png": self.also_png_var.get(),
        }

    def all_psd_layers_labels(self):
        """当前 PSD 全部图层的展示 label（与 logo_checks 的 key 一致）。"""
        if self.layer_index is not None:
            return [self.layer_labels[r.id] for r in self.layer_index.layers]
        return list(self.all_psd_layers)

    def _start(self):
        if self.running:
            return
        cfg = self._collect_cfg()
        if not cfg["psd_path"] or not os.path.exists(cfg["psd_path"]):
            messagebox.showerror("错误", "请先选择 PSD 模板。")
            return
        if not cfg["xlsx_path"] or not os.path.exists(cfg["xlsx_path"]):
            messagebox.showerror("错误", "请先选择 Excel 数据。")
            return
        if not cfg["out_dir"]:
            messagebox.showerror("错误", "请先选择输出目录。")
            return

        self.running = True
        self.stop_flag.clear()
        self.btn_run.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.btn_preview.config(state="disabled")
        self.progress["value"] = 0
        self._log_gui("开始批量生成...")

        self.worker = threading.Thread(target=self._worker_run, args=(cfg,), daemon=True)
        self.worker.start()

    def _worker_run(self, cfg):
        def prog(done, total):
            self.root.after(0, lambda: self._update_progress(done, total))
        def log(msg):
            self.root.after(0, lambda: self._log_gui(msg))
        run_batch(cfg, prog, log, self.stop_flag)
        self.root.after(0, self._on_finish)

    def _preview(self):
        if self.running:
            return
        cfg = self._collect_cfg()
        if not cfg["psd_path"] or not os.path.exists(cfg["psd_path"]) or \
           not cfg["xlsx_path"] or not os.path.exists(cfg["xlsx_path"]):
            messagebox.showerror("错误", "请先选择 PSD 与 Excel，并点击【加载】。")
            return
        tmp = os.path.join(os.path.dirname(os.path.abspath(__file__)), "preview_tmp")
        os.makedirs(tmp, exist_ok=True)
        cfg = dict(cfg)
        cfg["out_dir"] = tmp
        self._log_gui("生成预览（第1行数据）...")
        # 预览只做第一行：用 col 读取第1行（COM 初始化由 Session 负责）
        try:
            wb = openpyxl.load_workbook(cfg["xlsx_path"], data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            start = 1 if cfg["has_header"] else 0
            r = rows[start]
            store = str(r[cfg["col_store"]]).strip() if r[cfg["col_store"]] is not None else ""
            name = str(r[cfg["col_name"]]).strip()
            phone_v = r[cfg["col_phone"]]
            phone_s = str(int(phone_v)) if isinstance(phone_v, (int, float)) else str(phone_v).strip()
            role = None
            if cfg["col_role"] >= 0 and cfg["col_role"] < len(r) and r[cfg["col_role"]] is not None:
                role = str(r[cfg["col_role"]]).strip()

            with PhotoshopSession() as ps:
                app = ps.app
                doc0 = ps.open_document(cfg["psd_path"])
                time.sleep(0.6)

                # Stage 2：运行时 LayerIndex
                index = collect_layer_index(doc0)

                text_map = cfg.get("text_map", {})
                name_ref = text_map.get("姓名", "")
                phone_ref = text_map.get("电话", "")
                role_ref = text_map.get("销售顾问", "")

                d = None
                try:
                    d = ps.duplicate_document(doc0)
                    app.ActiveDocument = d
                    time.sleep(0.05)

                    # Stage 2：按 LayerRef 精确 resolve（duplicate 上重新定位）
                    if name_ref:
                        set_text_by_ref(d, index, name_ref, name, self._log_gui, label="姓名")
                    if phone_ref:
                        set_text_by_ref(d, index, phone_ref, phone_s, self._log_gui, label="电话")
                    if role_ref and role is not None:
                        set_text_by_ref(d, index, role_ref, role, self._log_gui, label="销售顾问")

                    mapped = cfg["store_logo_map"].get(store, "")
                    brand = brand_logos_for(cfg["logo_layers"], cfg["store_logo_map"])
                    brand_keys = {_ref_key(x) for x in brand}
                    mapped_key = _ref_key(mapped) if mapped else None
                    for ln in cfg["logo_layers"]:
                        if _ref_key(ln) in brand_keys:
                            set_visible_by_ref(d, index, ln, True, self._log_gui, label="品牌Logo")
                        else:
                            set_visible_by_ref(d, index, ln, (_ref_key(ln) == mapped_key),
                                               self._log_gui, label="门店Logo")
                    p = os.path.join(tmp, "preview.png")
                    export_doc(d, p, FMT_PNG)
                    self._log_gui(f"预览已生成：{p}")
                    try:
                        os.startfile(p)
                    except Exception:
                        pass
                finally:
                    # 无论预览成败，都关闭本工具创建的 duplicate，绝不波及用户文档
                    if d is not None:
                        try:
                            ps.close_owned_document(d)
                        except Exception as e:
                            self._log_gui(f"  警告：关闭预览副本失败：{e}")
        except Exception as e:
            self._log_gui(f"预览失败：{e}")

    def _stop(self):
        if self.running:
            self.stop_flag.set()
            self._log_gui("正在停止...")

    def _on_finish(self):
        self.running = False
        self.btn_run.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.btn_preview.config(state="normal")

    def _update_progress(self, done, total):
        if total > 0:
            self.progress["value"] = int(done / total * 100)
        self.progress_label.config(text=f"{done} / {total}")

    def _log_gui(self, msg):
        self.log.config(state="normal")
        self.log.insert("end", msg + "\n")
        self.log.see("end")
        self.log.config(state="disabled")

    # ---------------- 配置存取 ----------------
    def _save_config(self):
        cfg = self._collect_cfg()
        # 即使未加载，也保存路径与列设置
        try:
            with open(self.config_path, "w", encoding="utf-8") as f:
                json.dump(cfg, f, ensure_ascii=False, indent=2)
            self._log_gui(f"配置已保存：{self.config_path}")
        except Exception as e:
            self._log_gui(f"保存配置失败：{e}")

    def _load_config(self):
        if not os.path.exists(self.config_path):
            return
        try:
            with open(self.config_path, "r", encoding="utf-8") as f:
                c = json.load(f)
            self.psd_var.set(c.get("psd_path", ""))
            self.xlsx_var.set(c.get("xlsx_path", ""))
            self.out_var.set(c.get("out_dir", ""))
            self.header_var.set(c.get("has_header", False))
            self.col_store_var.set(chr(ord("A") + c.get("col_store", 0)))
            self.col_name_var.set(chr(ord("A") + c.get("col_name", 1)))
            self.col_phone_var.set(chr(ord("A") + c.get("col_phone", 3)))
            rv = c.get("col_role", 2)
            self.col_role_var.set("（不替换）" if rv < 0 else chr(ord("A") + rv))
            self.fmt_var.set(c.get("fmt", FMT_PNG))
            self.also_png_var.set(c.get("also_png", False))
            # 文字图层映射（加载后会在 _load 中按实际图层校正，这里仅预填）
            tm = c.get("text_map", {})
            self.tm_name_var.set(self._tm_display(tm.get("姓名", "")))
            self.tm_phone_var.set(self._tm_display(tm.get("电话", "")))
            self.tm_role_var.set(self._tm_display(tm.get("销售顾问", "")))
            # 保留上次的 logo 勾选与映射，供 _load 默认填充
            self.cfg = c
        except Exception:
            pass

    def _tm_display(self, v):
        """把配置值（LayerRef dict 或旧 name 字符串）转成 GUI 下拉显示 label（未加载 PSD 时原样返回）。"""
        if not v:
            return "（不替换）"
        if self.layer_index is None:
            return v
        ref = ref_from_config(v)
        if ref is None:
            return "（不替换）"
        if ref.id and ref.id in self._ref_by_id:
            return self.layer_labels[ref.id]
        m = self.layer_index.find_matching(ref.name)
        if len(m) == 1:
            return self.layer_labels[m[0].id]
        return "（不替换）"   # ambiguous / missing：不自动选


def main():
    root = tk.Tk()
    try:
        root.tk.call("encoding", "system", "utf-8")
    except Exception:
        pass
    App(root)

    def _on_exit():
        # Stage 1：不再依赖模块级 _PS_LAUNCHED_BY_US 做全局 Quit 判断。
        # Session 的 ownership 只属于各自实例；窗口退出时不做任何跨 Session 的
        # Photoshop Quit / 关闭用户文档动作（避免误伤用户正在使用的 Photoshop）。
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", _on_exit)
    root.mainloop()


if __name__ == "__main__":
    main()
